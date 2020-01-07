import os
from module1 import Date_collection
from openpyxl import Workbook
from openpyxl.styles import fonts, alignment, PatternFill, Border, Side
import datetime

def top_list_font(a):
    # 블럭 병합
    ws.merge_cells('A1:I3')
    # 행 높이 변경
    for i in range(0, 4):
        ws.row_dimensions[i].height = 10

    # 폰트 및 정렬 변경
    a.font = fonts.Font(bold=True, size=20)
    a.alignment = alignment.Alignment(horizontal='center', vertical='center')


def main_row():
    main_rows = ['0', '날짜', '미터기수입금', '영업횟수', '운행시간', '영업시간', '운행거리', '영업거리', '가동대수', '거리실차율']

    for i in range(1, 10):
        setr = ws.cell(row=4, column=i)
        setr.value = main_rows[i]
        setr.font = fonts.Font(bold=True)
        setr.alignment = alignment.Alignment(horizontal='center', vertical='center')
        setr.fill = PatternFill(fgColor='80FFFF', fill_type='solid', patternType='solid')  # fgcolor = 배경색

    # 열 너비 변경
    column_ap = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']
    for j in range(len(column_ap)):
        ws.column_dimensions[column_ap[j]].width = 12

#가동률 수치 시트 생성 및 값 넣기
def operating_sheet():
    values = ['가동률 수치', '시간대', '대수']
    rows = [4, 5, 6, 7, 8, 9, 10,11]
    time = ['10:00', '9:00', '8:00', '7:00', '6:00', '5:00', '4:00', '3:00']
    counting = [1, 0.875, 0.75, 0.625, 0.5, 0.375, 0.25, 0.125]
    operating = wb.create_sheet("가동률 수치")
    for i in range(2, 5):
        operating.cell(row=3, column=i, value=values[i - 2])
    for i in range(len(rows)):
        operating.cell(row=rows[i], column=3, value=time[i])
        operating.cell(row=rows[i], column=4, value=counting[i])

#함수 추가 클래스
class function_insert:
    #가동률 함수
    def operating(self):
        for i in range(len(car_list)):
            ws = wb[car_list[i]]
            for j in range(date_index, total_date_index):
                # func =
                oper = ws.cell(row=j, column=8)
                oper.value = '=IF($D$%d'%date_index + ':$D$%d'%(total_date_index-1)+'>\'가동률 수치\'!$C$4,\'가동률 수치\'!$D$4,IF($D$%d'%date_index + ':$D$%d'%(total_date_index-1)+'>\'가동률 수치\'!$C$5,\'가동률 수치\'!$D$5,IF($D$%d'%date_index + ':$D$%d'%(total_date_index-1)+'>\'가동률 수치\'!$C$6,\'가동률 수치\'!$D$6,IF($D$%d'%date_index + ':$D$%d'%(total_date_index-1)+'>\'가동률 수치\'!$C$7,\'가동률 수치\'!$D$7, IF($D$%d'%date_index + ':$D$%d'%(total_date_index-1)+'>\'가동률 수치\'!$C$8,\'가동률 수치\'!$D$8, IF($D$%d'%date_index + ':$D$%d'%(total_date_index-1)+'>\'가동률 수치\'!$C$9,\'가동률 수치\'!$D$9, IF($D$%d'%date_index + ':$D$%d'%(total_date_index-1)+'>\'가동률 수치\'!$C$10,\'가동률 수치\'!$D$10, IF($D$%d'%date_index + ':$D$%d'%(total_date_index-1)+'>\'가동률 수치\'!$C$11,\'가동률 수치\'!$D$11, 0))))))))'
    #가동대수 함수
    def load_factor(self):
        for i in range(len(car_list)):
            ws = wb[car_list[i]]
            for j in range(date_index, total_date_index):
                oper = ws.cell(row=j, column=9)
                oper.value ='=IF(ISERROR(G%d'%j+'/F%d'%j+'*100),,G%d'%j+'/F%d'%j+'*100)'
    #합계 함수
    def sum_line(self):
        col = ['0', 'A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']
        for i in range(len(car_list)):
            ws = wb[car_list[i]]
            for j in range(total_date_index, (total_date_index + 2)):
                for cols in range(2, 10):
                    ws.cell(row=total_date_index, column=cols,
                            value='=SUM(' + '%s' % col[cols] + '%d' % date_index + ':' + '%s' % col[cols] + '%d' % (
                                    total_date_index - 1) + ')')
                    
    #평균 함수
    def averge_line(self):
        col = ['0', 'A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']
        for i in range(len(car_list)):
            ws = wb[car_list[i]]
            for j in range(total_date_index, (total_date_index + 2)):
                for cols in range(2, 8):
                    ws.cell(row=(total_date_index+1), column=cols,
                            value='=%s' %col[cols]+'%d'%total_date_index+'/count(%s' %col[cols]+'%d'%date_index+':%s' %col[cols]+'%d)'%(total_date_index-1) )

#표시형식 바꾸기
def generator():
    for i in range(len(car_list)):
        ws = wb[car_list[i]]
        #미터수입금
        for j in range(date_index, (total_date_index + 2)):
                fo1 = ws.cell(row=j, column=2)
                fo1.number_format = '#,##0'
        #영업횟수
        for j in range(date_index, (total_date_index + 2)):
                fo1 = ws.cell(row=j, column=3)
                fo1.number_format = '0'
        #운행시간, 영업시간
        for j in range(date_index, (total_date_index + 2)):
            for cols in range(4, 6):
                fo1 = ws.cell(row=j, column=cols)
                fo1.number_format = '[hh]:mm'
        #운행거리, 영업거리
        for j in range(date_index, (total_date_index + 2)):
            for cols in range(6, 8):
                fo1 = ws.cell(row=j, column=cols)
                fo1.number_format = '#,##0.0'
        #가동대수
        for j in range(date_index, (total_date_index + 2)):
                fo1 = ws.cell(row=j, column=8)
                fo1.number_format = '#,##0.000'
        #거리실차율
        for j in range(date_index, (total_date_index + 2)):
                fo1 = ws.cell(row=j, column=9)
                fo1.number_format = '#,##0.0'
                

#폴더 이름과 파일이름 추출
folder_name = os.getcwd().split('\\')[-1]
print("folder_name : {}".format(folder_name))
path = './'
file_list = os.listdir(path)
file_list_py = [file for file in file_list if file.endswith(".xlsx")]

#워크북 생성
wb = Workbook()

#파일 이름으로 리스트 생성
car_list = []
car_list.extend(file_list_py[i].rstrip('.xlsx') for i in range(len(file_list_py)))
a = []
a.extend(wb.get_sheet_names())

thin_border = Border(top=Side(style='thin'), right=Side(style='thin'), bottom=Side(style='thin'),
                     left=Side(style='thin'))

#날짜계산
x = Date_collection()
plus_line = x.box_length()
main_line = 4
total_line = main_line + plus_line
date_index = 5
total_date_index = date_index + plus_line
adate = x.date_insert()

# 새로 생성된 엑셀 Sheet 지우기
for j in range(len(a)):
    if a[j].count('Sheet') != 0:
        del wb[a[j]]

# 메인
for i in range(len(car_list)):
    ws = wb.create_sheet(car_list[i])
    # 각 시트 활성화
    ws.cell(row=1, column=1, value=car_list[i])
    cf = ws['A1']
    top_list_font(cf)

    # 테두리 그리기
    cfs = ws['A%d' % main_line:'I%d' % (total_line + 2)]
    for row in cfs:
        for cell in row:
            cell.border = thin_border
            cell.alignment = alignment.Alignment(vertical='center')


    main_row()
    
    # 날자 삽입
    t = 0
    for row in range(date_index, total_date_index):
        dater = ws.cell(row=row, column=1)
        dater.value = adate[t]
        dater.number_format = 'yyyy-mm-dd'
        if t != total_date_index - date_index - 1:
            t = t + 1

    #합계및 평균 라인 색, 정렬
    for row in range(total_date_index, (total_date_index + 2)):
        for cols in range(1, 10):
            b = ws.cell(row=row, column=cols)
            b.fill = PatternFill(fill_type='solid', fgColor='FFC000', patternType='solid')
            b.alignment = alignment.Alignment(vertical='center', horizontal="general")


    for i in range((total_date_index), (total_date_index + 2)):
        ws.cell(row=(total_date_index), column=1, value="합계")
        ws.cell(row=(total_date_index + 1), column=1, value="평균")
        a = ws.cell(row=i, column=1)
        a.font = fonts.Font(bold=True)
        a.alignment = alignment.Alignment(horizontal='center')

operating_sheet()
function_insert().operating()
function_insert().load_factor()
function_insert().sum_line()
function_insert().averge_line()
generator()

# save
wb.save(folder_name + ' 통합 틀.xlsx')
